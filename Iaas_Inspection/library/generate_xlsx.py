#!/usr/bin/python
# -*- coding: utf-8 -*-


from __future__ import (absolute_import, division, print_function)

__metaclass__ = type
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

from ansible.module_utils.basic import AnsibleModule
from openpyxl import load_workbook
import json

class Tempalte(object):

    # eg: [(正常)'D38', (故障)'E38']
    platform_cell = {
        'platform_overview': {
	    'now_day': 'B3',
	    'engineer_name': 'E3',
	    'cluser_node_count': 'E4',
            'iaas_version': 'B4',
            'iaas_vm_sum': 'E6',
            'iaas_url': 'B5',
            'montior_url': 'E5',
            'ceph_cluster_overview': {
                'ceph_storage_total': 'B6',
                'ceph_storage_rate': 'C6',
                'ceph_health_status': ['D38', 'E38'],
                'ceph_osd_status': ['D37', 'E37'],
                # 'remark': 'I36'

            },
            'iaas_vcpus_overview': {
                'vcpus_core_total': 'B7',
                'vcpus_core_rate': 'C7'
            },
            'iaas_memory_overview': {
                'memory_total_gb': 'B8',
                'mepm_used_rate': 'C8'
            },
            'iaas_service_overview': {

                'ops_nova': ['D13', 'E13'],
                'ops_cinder': ['D14', 'E14'],
                'ops_neutron': ['D15', 'E15'],
                'ops_heat': ['D16', 'E16'],
                # 'remark': 'I13'
            },
            'mysql_cluster_status': ['D34', 'E34'],
            'rabbitmq_cluster_status': ['D35', 'E35']

        },
        'control_overview': {
            'control_ceph_mon': ['D36', 'E36'],
            'control_mysql': ['D17', 'E17'],
            'control_rabbitmq': ['D18', 'E18'],
            'control_horizon': ['D19', 'E19'],
            'control_keystone': ['D20', 'E20'],
            'control_memcached': ['D21', 'E21'],
            'control_neutron': ['D22', 'E22'],
            'control_nova': ['D23', 'E23'],
            'control_cinder': ['D24', 'E24'],
            'control_glance': ['D25', 'E25'],
            'control_heat': ['D26', 'E26'],
            'control_network': ['D27', 'E27'],
            'control_network_dhcp': ['D28', 'E28'],
            'control_keepalived': ['D29', 'E29'],
            'control_haproxy': ['D30', 'E30'],
            # 'remark': 'I19'

        },
        'compute_overview': {
            'cinder_volume': ['D33', 'E33'],
            'nova_compute': ['D32', 'E32'],
            'openvswitch_agent': ['D31', 'E31'],
            # 'remark': 'I31'
        }
    }

    # default row 2
    os_row = 2

    os_cell = {

        'default_ipv4': 'A{}',
        'hostname': 'B{}',
        'os_pretty_name': 'C{}',
        'uptime': 'D{}',
        'cpu_usedutilization': 'E{}',
        'mem_usedutilization': 'F{}',
        'size_usedutilization': 'G{}'
    }


class Generate_Xlsx(object):

    def __init__(self, module):

        self.module = module
        self.src_data = self.module.params['src_data']
        self.platform_sheet_name = self.module.params['platform_sheet_name']
        self.os_sheet_name = self.module.params['os_sheet_name']
        self.xlsx_template = self.module.params['xlsx_template']
        self.save_path = self.module.params['save_path']

        self.dict_data = self._json()
        self.xlsx_wb = self._xlsx()

    def __del__(self):
        self.xlsx_wb.save(self.save_path)
        self.xlsx_wb.close()

    def _json(self):
        try:
            with open(self.src_data, 'r') as f:
                d = f.read()
                return json.loads(d)
        except Exception as err:
            self.module.fail_json(err, **dict(stdout='<generate_xlsx> json loads data failure'))

    def _xlsx(self):
        try:
            self.wb = load_workbook(self.xlsx_template)
            return self.wb
        except Exception as err:
            self.module.fail_json(err, **dict(stdout='Failed to read xlsx file template'))

    def _write(self, coord, value):
        # eg: coord:A1 value: str unicode
        try:

            if isinstance(value, list):
                self.ws[coord[1]] = u'☑故障'
                return

            if isinstance(value, unicode):
                self.ws[coord] = value
                return 

            else:
                # eg: coord:A1
                if value:
                    self.ws[coord[0]] = u'☑正常'
                    return
                self.ws[coord[1]] = u'☑故障'
        except Exception as err:
            self.module.fail_json(err, **dict(stdout='Error writing xlsx file <coord: {} value: {}>'.format(coord, value)))



    def generate_platform(self, platform_dict):

        self.ws = self.xlsx_wb[self.platform_sheet_name]
        for viewkey, viewvalue in Tempalte.platform_cell.items():  # platform_overview control_overview  compute_overview
            for services, value_coord in viewvalue.items():
                if services in ('iaas_vm_sum'):
                    v = u'已经使用虚拟机个数__{}__'.format(platform_dict[viewkey][services])
                    self._write(value_coord, v)
                    continue
                if isinstance(value_coord, dict):
                    for service, row in value_coord.items():
                        if service in ('vcpus_core_rate','mepm_used_rate', 'ceph_storage_rate'):
                            v = u'使用率：{}'.format(platform_dict[viewkey][services][service])
                            self._write(row, v)
                            continue

                        if service in ('vcpus_core_total','memory_total_gb', 'ceph_storage_total'):
                            v = u'总容量: {}'.format(platform_dict[viewkey][services][service])
                            self._write(row, v)
                            continue

                        self._write(row, platform_dict[viewkey][services][service])
                else:
                    self._write(value_coord, platform_dict[viewkey][services])

    def generate_os(self, os_dict):
        self.ws = self.xlsx_wb[self.os_sheet_name]

        for host_ip, host_info in os_dict.items():
            for key, value in Tempalte.os_cell.items():
                if os_dict[host_ip].get('unreachable', False):
                    self._write('A{}'.format(Tempalte.os_row), host_ip)
                    break
                self._write(value.format(Tempalte.os_row), host_info[key])
            Tempalte.os_row += 1


def main():
    module = AnsibleModule(
        argument_spec=dict(
            src_data=dict(required=True, type='str'),
            platform_sheet_name=dict(default='平台巡检单', type='str'),
            save_path=dict(default=True, type='str'),
            xlsx_template=dict(default=True, type='str'),
            os_sheet_name=dict(default='主机负载情况', type='str')
        ),
        supports_check_mode = True,
    )

    generate_tool = Generate_Xlsx(module)

    if generate_tool.dict_data['platform']:
       generate_tool.generate_platform(generate_tool.dict_data['platform'])

    if generate_tool.dict_data['os']:
        generate_tool.generate_os(generate_tool.dict_data['os'])
    generate_tool.module.exit_json(**dict(rc=0, stdout='The inspection report is generated successfully', changed=True))


if __name__ == '__main__':
    main()
